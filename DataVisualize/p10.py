import pandas as pd
import os
from openpyxl import load_workbook

def extract_wastewater_data(sheet, months_offset):
    """
    从废水sheet提取数据
    months_offset: 季度对应的月份偏移量（1, 4, 7, 10）
    """
    monthly_data = {}
    
    # 查找污染物数据起始位置
    data_start_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "排放口" and sheet.cell(row=row, column=2).value == "污染物名称":
            data_start_row = row + 1
            break
    
    if not data_start_row:
        return monthly_data
    
    # 污染物列表（按ALLINONE顺序）
    pollutants = [
        "COD", "氨氮", "六价铬", "石油类", "悬浮物", "氟化物", 
        "铅", "镍", "铬", "BOD5", "总氮", "总磷", "Ph"
    ]
    
    # 提取每个污染物3个月的数据
    for i, pollutant in enumerate(pollutants):
        row = data_start_row + i
        
        # 获取3个月的实际排放量数据
        # 根据文件结构，排放量数据在以下列：
        # 1月：F列（第6列），2月：J列（第10列），3月：N列（第14列）
        month1_value = sheet.cell(row=row, column=6).value  # 1月排放量
        month2_value = sheet.cell(row=row, column=10).value  # 2月排放量
        month3_value = sheet.cell(row=row, column=14).value  # 3月排放量
        
        # 对于pH值，从浓度列获取
        if pollutant == "Ph":
            month1_value = sheet.cell(row=row, column=3).value  # 1月浓度
            month2_value = sheet.cell(row=row, column=7).value  # 2月浓度
            month3_value = sheet.cell(row=row, column=11).value  # 3月浓度
        
        monthly_data[pollutant] = {
            months_offset: month1_value if month1_value is not None else 0,
            months_offset + 1: month2_value if month2_value is not None else 0,
            months_offset + 2: month3_value if month3_value is not None else 0
        }
    
    return monthly_data

def extract_exhaust_data(sheet, quarter_num):
    """从废气sheet提取季度总量数据"""
    data = {}
    
    # 污染物及其总计行的映射
    pollutant_totals = {}
    
    # 扫描整个表格找到污染物名称和总计行
    current_pollutant = None
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=2).value
        
        if cell_value in ["颗粒物", "NOX", "SO2", "VOC", "氟化物", "氯化氢"]:
            current_pollutant = cell_value
        elif sheet.cell(row=row, column=1).value == "总计" and current_pollutant:
            # 找到总计行，获取O列（第15列）的季度合计值
            quarter_total = sheet.cell(row=row, column=15).value
            if quarter_total is not None:
                data[current_pollutant] = quarter_total
            current_pollutant = None
    
    return data

def process_quarter_file(file_path, year, quarter_num):
    """处理单个季度的文件"""
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # 废水数据
        wastewater_sheet = wb['废水']
        # 计算季度对应的起始月份
        month_offset = (quarter_num - 1) * 3 + 1
        wastewater_data = extract_wastewater_data(wastewater_sheet, month_offset)
        
        # 废气数据
        exhaust_sheet = wb['废气']
        exhaust_data = extract_exhaust_data(exhaust_sheet, quarter_num)
        
        return {
            'year': year,
            'quarter': quarter_num,
            'wastewater': wastewater_data,
            'exhaust': exhaust_data
        }
        
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return None

def main():
    # 初始化结果字典
    years = [2023, 2024, 2025]
    
    # 污染物列表（按ALLINONE顺序）
    pollutants = [
        "COD", "氨氮", "六价铬", "石油类", "悬浮物", "氟化物", 
        "铅", "镍", "铬", "BOD5", "总氮", "总磷", "Ph",
        "颗粒物", "NOX", "SO2", "VOC", "氟化物", "氯化氢"
    ]
    
    # 初始化结果字典
    results = {}
    for pollutant in pollutants:
        results[pollutant] = {}
        for year in years:
            for month in range(1, 13):
                results[pollutant][f"{year}年{month}月"] = ""
    
    # 处理所有文件
    for year in years:
        for quarter in range(1, 5):  # 4个季度
            file_name = f"{year}年第{quarter}季度废水废气实际排放量计算结果.xlsx"
            
            if os.path.exists(file_name):
                print(f"正在处理: {file_name}")
                data = process_quarter_file(file_name, year, quarter)
                
                if data:
                    # 处理废水数据
                    for pollutant, month_values in data['wastewater'].items():
                        if pollutant in results:
                            for month_offset, value in month_values.items():
                                month_key = f"{year}年{month_offset}月"
                                if month_key in results[pollutant]:
                                    results[pollutant][month_key] = value if value is not None else ""
                    
                    # 处理废气数据 - 季度总量放到每个月的单元格
                    # 注意：根据ALLINONE示例，废气数据是季度总量，不是月平均值
                    for pollutant, quarter_total in data['exhaust'].items():
                        if pollutant in results:
                            # 计算季度对应的月份
                            start_month = (quarter - 1) * 3 + 1
                            months = [start_month, start_month + 1, start_month + 2]
                            
                            # 将季度总量放入第一个月，其他月留空
                            first_month_key = f"{year}年{start_month}月"
                            results[pollutant][first_month_key] = quarter_total if quarter_total is not None else ""
                            
                            # 其他两个月留空（根据ALLINONE示例）
                            for month in months[1:]:
                                month_key = f"{year}年{month}月"
                                results[pollutant][month_key] = ""
            else:
                print(f"文件不存在: {file_name}")
    
    # 创建DataFrame
    # 准备列名：污染物名称 + 所有年月
    columns = ['污染物名称']
    for year in years:
        for month in range(1, 13):
            columns.append(f"{year}年{month}月")
    
    # 准备数据行
    data_rows = []
    for pollutant in pollutants:
        row = {'污染物名称': pollutant}
        for year in years:
            for month in range(1, 13):
                month_key = f"{year}年{month}月"
                row[month_key] = results[pollutant].get(month_key, "")
        data_rows.append(row)
    
    # 创建DataFrame
    df = pd.DataFrame(data_rows, columns=columns)
    
    # 保存到Excel
    output_file = "污染物排放量汇总_三年数据.xlsx"
    df.to_excel(output_file, index=False, sheet_name='汇总')
    
    print(f"\n数据处理完成！结果已保存到: {output_file}")
    print(f"共处理了 {len(pollutants)} 种污染物，{len(years)*12} 个月份的数据")
    
    # 显示数据预览
    print("\n数据预览:")
    print(df.head(20))
    
    return df

if __name__ == "__main__":
    # 如果需要查看特定污染物的数据
    df = main()
    
    # 示例：查看COD的所有数据
    print("\nCOD数据示例:")
    cod_row = df[df['污染物名称'] == 'COD']
    print(cod_row)